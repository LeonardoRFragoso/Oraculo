import csv
from openpyxl import load_workbook

def converter_xlsx_para_csv(arquivo_entrada, arquivo_saida):
    # Carregar a planilha
    wb = load_workbook(filename=arquivo_entrada, read_only=True)
    ws = wb.active  # Primeira aba ativa

    # Abrir o arquivo CSV para escrita
    with open(arquivo_saida, mode='w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    print(f'Conversão concluída: {arquivo_saida}')


if __name__ == "__main__":
    # Substitua os caminhos conforme necessário
    caminho_entrada = "Dados_Consolidados.xlsx"
    caminho_saida = "Dados_Consolidados.csv"
    converter_xlsx_para_csv(caminho_entrada, caminho_saida)
